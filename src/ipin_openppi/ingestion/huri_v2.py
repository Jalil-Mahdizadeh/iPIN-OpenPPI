"""HuRI parser revision with deterministic legacy/workbook support.

The interaction evidence and public pair-view logic remains in :mod:`huri`.
This revision replaces only supplementary parsing so every publication table is
covered, while contact and fusion-interference annotations remain explicitly
unauthorized for label construction.
"""

from __future__ import annotations

from collections import Counter
import csv
from datetime import date, datetime, time
import io
import math
from pathlib import Path
import re
from typing import Any
from urllib.parse import quote
import zipfile

import openpyxl
import xlrd

from .common import ParquetBatchWriter, canonical_json, stable_id, strip_version
from .context import ParsingContext
from .huri import _parse_mitab, _parse_pair_views, _raw_bool


_SUPPLEMENT_MEMBER_RE = re.compile(
    r"Supplementary Table (\d+)\.(txt|xls|xlsx)$", re.IGNORECASE
)
_EXPECTED_TABLE_NUMBERS = set(range(1, 30))

_TABLE_10_SHEETS = {"info", "HuRI", "bioplex", "cofrac", "qubic", "litBM"}
_TABLE_10_DATA_SHEETS = _TABLE_10_SHEETS - {"info"}
_TABLE_15_SHEETS = {"legend", "20190418_fusion_interference"}
_TABLE_21_HEADER_ROWS = {
    "Sheet2": 1,
    "Lethality": 3,
    "LoF intolerance": 3,
    "Fitness Effect": 3,
    "Age": 3,
    "Publication": 3,
    "Expression (GTEx)": 3,
    "Expression (BioPlex)": 3,
    "Expression (QUBIC)": 3,
}

_EXPECTED_HEADERS: dict[tuple[int, str], list[str]] = {
    **{
        (10, sheet): ["protein1", "protein2", "in_contact"]
        for sheet in _TABLE_10_DATA_SHEETS
    },
    (15, "20190418_fusion_interference"): [
        "DBD_fused",
        "AD_fused",
        "d_N1_iface",
        "d_C1_iface",
        "d_N2_iface",
        "d_C2_iface",
        "found_v1",
        "found_v2",
        "found_v3",
    ],
    (21, "Sheet2"): [
        "Category",
        "Number of genes",
        "Source",
        "Selection criteria",
    ],
    (21, "Lethality"): ["Index", "ENSG accession", "Gene symbol"],
    (21, "LoF intolerance"): [
        "Index",
        "ENSG accession",
        "Gene symbol",
        "pLI",
    ],
    (21, "Fitness Effect"): [
        "Index",
        "ENSG accession",
        "Gene symbol",
        "Fitness effect",
    ],
    (21, "Age"): ["Index", "ENSG accession", "Gene symbol", "Age"],
    (21, "Publication"): [
        "Index",
        "ENSG accession",
        "Gene symbol",
        "Number of publications",
    ],
    (21, "Expression (GTEx)"): [
        "Index",
        "ENSG accession",
        "Gene symbol",
        "Median expression",
    ],
    (21, "Expression (BioPlex)"): [
        "Index",
        "ENSG accession",
        "Gene symbol",
        "Expression",
    ],
    (21, "Expression (QUBIC)"): [
        "Index",
        "ENSG accession",
        "Gene symbol",
        "Copy number",
    ],
    (29, "Sheet1"): ["Figure", "Panel", "n"],
}

_XLRD_TYPE_NAMES = {
    xlrd.XL_CELL_EMPTY: "empty",
    xlrd.XL_CELL_TEXT: "text",
    xlrd.XL_CELL_NUMBER: "number",
    xlrd.XL_CELL_DATE: "date",
    xlrd.XL_CELL_BOOLEAN: "boolean",
    xlrd.XL_CELL_ERROR: "error",
    xlrd.XL_CELL_BLANK: "blank",
}


def _safe_json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, float) and not math.isfinite(value):
        raise ValueError(f"Workbook contains non-finite numeric value: {value!r}")
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _xls_cell(book: xlrd.book.Book, cell: xlrd.sheet.Cell) -> tuple[Any, str]:
    cell_type = _XLRD_TYPE_NAMES.get(cell.ctype, f"unknown:{cell.ctype}")
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        value: Any = None
    elif cell.ctype == xlrd.XL_CELL_DATE:
        value = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode).isoformat()
    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
        value = bool(cell.value)
    elif cell.ctype == xlrd.XL_CELL_ERROR:
        value = xlrd.biffh.error_text_from_code.get(cell.value, f"error:{cell.value}")
    else:
        value = _safe_json_value(cell.value)
    return value, cell_type


def _xlsx_cell(cell: Any) -> tuple[Any, str]:
    return _safe_json_value(cell.value), str(cell.data_type)


def _headers(values: list[Any], *, member: str, sheet: str, row: int) -> list[str]:
    headers = [str(value).strip() if value is not None else "" for value in values]
    if any(not value for value in headers):
        raise ValueError(f"Blank workbook header in {member} sheet {sheet!r} row {row}")
    if len(set(headers)) != len(headers):
        raise ValueError(
            f"Duplicate workbook header in {member} sheet {sheet!r} row {row}"
        )
    return headers


def _assert_expected_headers(
    table_number: int, member: str, sheet: str, headers: list[str]
) -> None:
    expected = _EXPECTED_HEADERS.get((table_number, sheet))
    if expected is not None and headers != expected:
        raise ValueError(
            f"Unexpected header in {member} sheet {sheet!r}: {headers!r} != {expected!r}"
        )


def _sheet_locator(member: str, sheet: str, physical_row: int) -> str:
    return f"zip:{member}#sheet:{quote(sheet, safe='')}#row:{physical_row}"


def _append_generic_workbook_row(
    *,
    writer: ParquetBatchWriter,
    context: ParsingContext,
    cfg: dict[str, Any],
    asset: Any,
    table_number: int,
    member: str,
    sheet: str,
    physical_row: int,
    record_kind: str,
    values: list[Any],
    cell_types: list[str],
    headers: list[str] | None,
    header_row: int | None,
) -> str:
    raw_locator = _sheet_locator(member, sheet, physical_row)
    if headers is None:
        payload: dict[str, Any] = {
            "record_kind": record_kind,
            "values": values,
            "cell_types": cell_types,
            "physical_row": physical_row,
        }
    else:
        payload = {
            "record_kind": record_kind,
            "fields": dict(zip(headers, values, strict=True)),
            "cell_types": dict(zip(headers, cell_types, strict=True)),
            "header_row": header_row,
            "physical_row": physical_row,
        }
    writer.append(
        {
            "staging_record_id": stable_id(
                "huri-supp-workbook-row",
                asset.sha256,
                member,
                sheet,
                physical_row,
            ),
            "source_key": "huri",
            "source_dataset": f"huri_supplement_table_{table_number}:{sheet}",
            "source_release": str(cfg["source_release"]),
            "source_member": member,
            "source_record_ordinal": physical_row,
            "raw_file_path": asset.relative_path,
            "raw_file_sha256": asset.sha256,
            "raw_locator": raw_locator,
            "fields_json": canonical_json(payload),
            "redistribution_tier": str(cfg["supplement_redistribution_tier"]),
        }
    )
    return raw_locator


def _strict_bool(value: Any, *, locator: str, field: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        parsed = _raw_bool(value)
        if parsed is not None:
            return parsed
    raise ValueError(f"Expected source boolean for {field} at {locator}: {value!r}")


def _required_text(value: Any, *, locator: str, field: str) -> str:
    if isinstance(value, str) and value.strip():
        return value.strip()
    raise ValueError(f"Expected non-empty text for {field} at {locator}: {value!r}")


def _optional_nonnegative_float(
    value: Any, *, locator: str, field: str
) -> tuple[float | None, str | None]:
    if value is None or (
        isinstance(value, str) and value.strip().casefold() in {"", "na", "n/a"}
    ):
        return None, "source_reported_na"
    if isinstance(value, bool):
        raise ValueError(f"Boolean in numeric field {field} at {locator}")
    try:
        parsed = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(
            f"Invalid numeric field {field} at {locator}: {value!r}"
        ) from exc
    if not math.isfinite(parsed) or parsed < 0:
        raise ValueError(
            f"Invalid non-negative distance {field} at {locator}: {value!r}"
        )
    return parsed, None


def _append_contact_row(
    *,
    writer: ParquetBatchWriter,
    cfg: dict[str, Any],
    asset: Any,
    member: str,
    sheet: str,
    physical_row: int,
    raw_locator: str,
    fields: dict[str, Any],
) -> None:
    protein_a = _required_text(
        fields["protein1"], locator=raw_locator, field="protein1"
    )
    protein_b = _required_text(
        fields["protein2"], locator=raw_locator, field="protein2"
    )
    in_contact = _strict_bool(
        fields["in_contact"], locator=raw_locator, field="in_contact"
    )
    writer.append(
        {
            "contact_record_id": stable_id(
                "huri-contact", asset.sha256, member, sheet, physical_row
            ),
            "source_release": str(cfg["source_release"]),
            "source_dataset": sheet,
            "protein_a_uniprot": protein_a,
            "protein_b_uniprot": protein_b,
            "unordered_pair_id": stable_id(
                "uniprot-pair", *sorted((protein_a, protein_b))
            ),
            "in_contact": in_contact,
            "label_authorized": False,
            "raw_file_path": asset.relative_path,
            "raw_file_sha256": asset.sha256,
            "raw_locator": raw_locator,
            "fields_json": canonical_json(fields),
            "redistribution_tier": str(cfg["supplement_redistribution_tier"]),
        }
    )


def _append_fusion_row(
    *,
    writer: ParquetBatchWriter,
    cfg: dict[str, Any],
    asset: Any,
    member: str,
    sheet: str,
    physical_row: int,
    raw_locator: str,
    fields: dict[str, Any],
) -> None:
    dbd = _required_text(fields["DBD_fused"], locator=raw_locator, field="DBD_fused")
    ad = _required_text(fields["AD_fused"], locator=raw_locator, field="AD_fused")
    distance_fields = {
        "distance_n1_interface": "d_N1_iface",
        "distance_c1_interface": "d_C1_iface",
        "distance_n2_interface": "d_N2_iface",
        "distance_c2_interface": "d_C2_iface",
    }
    distances: dict[str, float | None] = {}
    missingness: dict[str, str] = {}
    for target, source in distance_fields.items():
        parsed, missing_reason = _optional_nonnegative_float(
            fields[source], locator=raw_locator, field=source
        )
        distances[target] = parsed
        if missing_reason:
            missingness[target] = missing_reason
    writer.append(
        {
            "fusion_record_id": stable_id(
                "huri-fusion", asset.sha256, member, sheet, physical_row
            ),
            "source_release": str(cfg["source_release"]),
            "dbd_fused_uniprot": dbd,
            "ad_fused_uniprot": ad,
            "unordered_pair_id": stable_id("uniprot-pair", *sorted((dbd, ad))),
            "self_pair": dbd == ad,
            **distances,
            "found_v1": _strict_bool(
                fields["found_v1"], locator=raw_locator, field="found_v1"
            ),
            "found_v2": _strict_bool(
                fields["found_v2"], locator=raw_locator, field="found_v2"
            ),
            "found_v3": _strict_bool(
                fields["found_v3"], locator=raw_locator, field="found_v3"
            ),
            "label_authorized": False,
            "raw_file_path": asset.relative_path,
            "raw_file_sha256": asset.sha256,
            "raw_locator": raw_locator,
            "fields_json": canonical_json(fields),
            "missingness_json": canonical_json(missingness),
            "redistribution_tier": str(cfg["supplement_redistribution_tier"]),
        }
    )


def _parse_text_member(
    *,
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    table_number: int,
    context: ParsingContext,
    cfg: dict[str, Any],
    asset: Any,
    generic_writer: ParquetBatchWriter,
    orf_writer: ParquetBatchWriter,
    space_writer: ParquetBatchWriter,
) -> dict[str, Any]:
    dataset = f"huri_supplement_table_{table_number}"
    with archive.open(info) as binary:
        text = io.TextIOWrapper(binary, encoding="utf-8-sig", newline="")
        reader = csv.DictReader(text, delimiter="\t")
        if reader.fieldnames is None:
            raise ValueError(f"Supplement table lacks header: {info.filename}")
        if len(set(reader.fieldnames)) != len(reader.fieldnames):
            raise ValueError(f"Supplement table has duplicate headers: {info.filename}")
        row_count = 0
        field_value_counts: dict[str, Counter[str]] = {
            name: Counter() for name in reader.fieldnames
        }
        for row_ordinal, row in enumerate(reader, start=1):
            if None in row or any(value is None for value in row.values()):
                raise ValueError(
                    f"Malformed supplementary row in {info.filename} record {row_ordinal}"
                )
            fields = {str(key): str(value) for key, value in row.items()}
            raw_locator = f"zip:{info.filename}#data_row:{row_ordinal}"
            generic_writer.append(
                {
                    "staging_record_id": stable_id(
                        "huri-supp-row", asset.sha256, info.filename, row_ordinal
                    ),
                    "source_key": "huri",
                    "source_dataset": dataset,
                    "source_release": str(cfg["source_release"]),
                    "source_member": info.filename,
                    "source_record_ordinal": row_ordinal,
                    "raw_file_path": asset.relative_path,
                    "raw_file_sha256": asset.sha256,
                    "raw_locator": raw_locator,
                    "fields_json": canonical_json(fields),
                    "redistribution_tier": str(cfg["supplement_redistribution_tier"]),
                }
            )
            if table_number == 1:
                space_writer.append(
                    {
                        "space_record_id": stable_id(
                            "huri-space", asset.sha256, row_ordinal
                        ),
                        "ensembl_gene_id": strip_version(fields["Ensembl_gene_id"]),
                        "in_space_3": _raw_bool(fields["in_space_3"]),
                        "in_gtex": _raw_bool(fields["in_GTEx"]),
                        "in_hpa": _raw_bool(fields["in_HPA"]),
                        "in_fantom": _raw_bool(fields["in_FANTOM"]),
                        "raw_file_path": asset.relative_path,
                        "raw_locator": raw_locator,
                        "redistribution_tier": str(
                            cfg["supplement_redistribution_tier"]
                        ),
                    }
                )
            elif table_number == 2:
                field_names = {
                    "ensembl_transcript_id": "ensembl_transcript_id",
                    "ensembl_protein_id": "ensembl_protein_id",
                    "ensembl_gene_id": "ensembl_gene_id",
                    "gene_symbol": "symbol",
                }
                missingness = {
                    target: "not_reported"
                    for target, source in field_names.items()
                    if not fields.get(source)
                }
                orf_writer.append(
                    {
                        "orf_mapping_id": stable_id(
                            "huri-orf-map", asset.sha256, row_ordinal
                        ),
                        "orf_id": fields["orf_id"],
                        "ensembl_transcript_id": fields["ensembl_transcript_id"]
                        or None,
                        "ensembl_protein_id": fields["ensembl_protein_id"] or None,
                        "ensembl_gene_id": fields["ensembl_gene_id"] or None,
                        "gene_symbol": fields["symbol"] or None,
                        "raw_file_path": asset.relative_path,
                        "raw_locator": raw_locator,
                        "redistribution_tier": str(
                            cfg["supplement_redistribution_tier"]
                        ),
                        "missingness_json": canonical_json(missingness),
                    }
                )
            row_count += 1
            if table_number in {1, 9, 11, 16}:
                for field, value in fields.items():
                    field_value_counts[field][value] += 1
    return {
        "member": info.filename,
        "format": "tab_delimited_text",
        "rows": row_count,
        "columns": reader.fieldnames,
        "selected_value_counts": {
            field: dict(counter.most_common(30))
            for field, counter in field_value_counts.items()
            if len(counter) <= 30
        },
    }


def _parse_xls_member(
    *,
    payload: bytes,
    member: str,
    table_number: int,
    cfg: dict[str, Any],
    asset: Any,
    generic_writer: ParquetBatchWriter,
    contact_writer: ParquetBatchWriter,
    fusion_writer: ParquetBatchWriter,
) -> dict[str, Any]:
    book = xlrd.open_workbook(file_contents=payload, on_demand=True)
    observed_sheets = set(book.sheet_names())
    expected_sheets = {
        10: _TABLE_10_SHEETS,
        15: _TABLE_15_SHEETS,
        29: {"Sheet1"},
    }.get(table_number)
    if expected_sheets is None or observed_sheets != expected_sheets:
        raise ValueError(
            f"Unexpected sheets in {member}: {sorted(observed_sheets)} != "
            f"{sorted(expected_sheets or set())}"
        )
    sheet_stats: dict[str, Any] = {}
    try:
        for sheet in book.sheets():
            header_row = (
                1
                if sheet.name in _TABLE_10_DATA_SHEETS
                or (table_number, sheet.name) in _EXPECTED_HEADERS
                else None
            )
            headers: list[str] | None = None
            data_rows = 0
            metadata_rows = 0
            error_cells = 0
            for zero_based_row in range(sheet.nrows):
                physical_row = zero_based_row + 1
                converted = [
                    _xls_cell(book, cell) for cell in sheet.row(zero_based_row)
                ]
                values = [value for value, _ in converted]
                cell_types = [cell_type for _, cell_type in converted]
                error_cells += sum(cell_type == "error" for cell_type in cell_types)
                if physical_row == header_row:
                    headers = _headers(
                        values, member=member, sheet=sheet.name, row=physical_row
                    )
                    _assert_expected_headers(table_number, member, sheet.name, headers)
                    continue
                is_data = header_row is not None and physical_row > header_row
                if is_data and headers is None:
                    raise RuntimeError(
                        f"Header state lost for {member} sheet {sheet.name}"
                    )
                raw_locator = _append_generic_workbook_row(
                    writer=generic_writer,
                    context=None,  # retained for a stable call signature
                    cfg=cfg,
                    asset=asset,
                    table_number=table_number,
                    member=member,
                    sheet=sheet.name,
                    physical_row=physical_row,
                    record_kind=(
                        "workbook_data_row" if is_data else "workbook_metadata_row"
                    ),
                    values=values,
                    cell_types=cell_types,
                    headers=headers if is_data else None,
                    header_row=header_row,
                )
                if is_data:
                    fields = dict(zip(headers or [], values, strict=True))
                    if table_number == 10:
                        _append_contact_row(
                            writer=contact_writer,
                            cfg=cfg,
                            asset=asset,
                            member=member,
                            sheet=sheet.name,
                            physical_row=physical_row,
                            raw_locator=raw_locator,
                            fields=fields,
                        )
                    elif table_number == 15:
                        _append_fusion_row(
                            writer=fusion_writer,
                            cfg=cfg,
                            asset=asset,
                            member=member,
                            sheet=sheet.name,
                            physical_row=physical_row,
                            raw_locator=raw_locator,
                            fields=fields,
                        )
                    data_rows += 1
                else:
                    metadata_rows += 1
            sheet_stats[sheet.name] = {
                "rows": sheet.nrows,
                "columns": sheet.ncols,
                "header_row": header_row,
                "headers": headers,
                "data_rows": data_rows,
                "metadata_rows": metadata_rows,
                "source_error_cells": error_cells,
            }
    finally:
        book.release_resources()
    return {"member": member, "format": "xls", "sheets": sheet_stats}


def _parse_xlsx_member(
    *,
    payload: bytes,
    member: str,
    table_number: int,
    cfg: dict[str, Any],
    asset: Any,
    generic_writer: ParquetBatchWriter,
) -> dict[str, Any]:
    if table_number != 21:
        raise ValueError(f"Unexpected XLSX supplementary table: {member}")
    book = openpyxl.load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    observed_sheets = set(book.sheetnames)
    expected_sheets = set(_TABLE_21_HEADER_ROWS)
    if observed_sheets != expected_sheets:
        book.close()
        raise ValueError(
            f"Unexpected sheets in {member}: {sorted(observed_sheets)} != "
            f"{sorted(expected_sheets)}"
        )
    sheet_stats: dict[str, Any] = {}
    try:
        for sheet in book.worksheets:
            header_row = _TABLE_21_HEADER_ROWS[sheet.title]
            headers: list[str] | None = None
            data_rows = 0
            metadata_rows = 0
            error_cells = 0
            for physical_row, cells in enumerate(sheet.iter_rows(), start=1):
                converted = [_xlsx_cell(cell) for cell in cells]
                values = [value for value, _ in converted]
                cell_types = [cell_type for _, cell_type in converted]
                error_cells += sum(cell_type == "e" for cell_type in cell_types)
                if physical_row == header_row:
                    headers = _headers(
                        values, member=member, sheet=sheet.title, row=physical_row
                    )
                    _assert_expected_headers(table_number, member, sheet.title, headers)
                    continue
                is_data = physical_row > header_row
                if is_data and headers is None:
                    raise RuntimeError(
                        f"Header state lost for {member} sheet {sheet.title}"
                    )
                _append_generic_workbook_row(
                    writer=generic_writer,
                    context=None,
                    cfg=cfg,
                    asset=asset,
                    table_number=table_number,
                    member=member,
                    sheet=sheet.title,
                    physical_row=physical_row,
                    record_kind=(
                        "workbook_data_row" if is_data else "workbook_metadata_row"
                    ),
                    values=values,
                    cell_types=cell_types,
                    headers=headers if is_data else None,
                    header_row=header_row,
                )
                if is_data:
                    data_rows += 1
                else:
                    metadata_rows += 1
            sheet_stats[sheet.title] = {
                "rows": sheet.max_row,
                "columns": sheet.max_column,
                "header_row": header_row,
                "headers": headers,
                "data_rows": data_rows,
                "metadata_rows": metadata_rows,
                "source_error_cells": error_cells,
            }
    finally:
        book.close()
    return {"member": member, "format": "xlsx", "sheets": sheet_stats}


def _parse_supplementary_tables(
    context: ParsingContext, output_root: Path, cfg: dict[str, Any]
) -> tuple[
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
    dict[str, Any],
]:
    asset = context.asset(str(cfg["supplementary_asset_id"]))
    generic_writer = ParquetBatchWriter(
        output_root / "supplementary_raw_tabular_records",
        context.staging_contract,
        "raw_tabular_records",
        **context.writer_kwargs(),
    )
    orf_writer = ParquetBatchWriter(
        output_root / "huri_orf_mappings",
        context.evidence_contract,
        "huri_orf_mappings",
        **context.writer_kwargs(),
    )
    space_writer = ParquetBatchWriter(
        output_root / "huri_space_membership",
        context.evidence_contract,
        "huri_space_membership",
        **context.writer_kwargs(),
    )
    contact_writer = ParquetBatchWriter(
        output_root / "huri_structural_contact_annotations",
        context.staging_contract,
        "huri_structural_contact_annotations",
        **context.writer_kwargs(),
    )
    fusion_writer = ParquetBatchWriter(
        output_root / "huri_fusion_interference",
        context.staging_contract,
        "huri_fusion_interference",
        **context.writer_kwargs(),
    )
    table_stats: dict[str, Any] = {}
    unparsed_members: list[dict[str, Any]] = []
    ignored_members: list[dict[str, Any]] = []
    archive_inventory: list[dict[str, Any]] = []
    table_counts: Counter[int] = Counter()
    with (
        zipfile.ZipFile(asset.path) as archive,
        generic_writer,
        orf_writer,
        space_writer,
        contact_writer,
        fusion_writer,
    ):
        for info in archive.infolist():
            archive_inventory.append(
                {
                    "member": info.filename,
                    "bytes": info.file_size,
                    "compressed_bytes": info.compress_size,
                    "is_directory": info.is_dir(),
                }
            )
            if info.is_dir():
                ignored_members.append(
                    {"member": info.filename, "reason": "archive_directory"}
                )
                continue
            if info.filename.startswith("__MACOSX/"):
                ignored_members.append(
                    {"member": info.filename, "reason": "macos_resource_fork_metadata"}
                )
                continue
            match = _SUPPLEMENT_MEMBER_RE.search(info.filename)
            if not match:
                unparsed_members.append(
                    {"member": info.filename, "reason": "unsupported_member_name"}
                )
                continue
            table_number = int(match.group(1))
            extension = match.group(2).casefold()
            table_counts[table_number] += 1
            if extension == "txt":
                table_stats[str(table_number)] = _parse_text_member(
                    archive=archive,
                    info=info,
                    table_number=table_number,
                    context=context,
                    cfg=cfg,
                    asset=asset,
                    generic_writer=generic_writer,
                    orf_writer=orf_writer,
                    space_writer=space_writer,
                )
            elif extension == "xls":
                table_stats[str(table_number)] = _parse_xls_member(
                    payload=archive.read(info),
                    member=info.filename,
                    table_number=table_number,
                    cfg=cfg,
                    asset=asset,
                    generic_writer=generic_writer,
                    contact_writer=contact_writer,
                    fusion_writer=fusion_writer,
                )
            elif extension == "xlsx":
                table_stats[str(table_number)] = _parse_xlsx_member(
                    payload=archive.read(info),
                    member=info.filename,
                    table_number=table_number,
                    cfg=cfg,
                    asset=asset,
                    generic_writer=generic_writer,
                )
            else:  # pragma: no cover - guarded by the regular expression
                raise AssertionError(extension)
    if unparsed_members:
        raise ValueError(f"Unparsed HuRI supplementary members: {unparsed_members!r}")
    if set(table_counts) != _EXPECTED_TABLE_NUMBERS or any(
        count != 1 for count in table_counts.values()
    ):
        raise ValueError(
            "HuRI supplementary archive does not contain exactly one table for each "
            f"number 1-29: {dict(sorted(table_counts.items()))}"
        )
    return (
        generic_writer.summary(),
        orf_writer.summary(),
        space_writer.summary(),
        contact_writer.summary(),
        fusion_writer.summary(),
        {
            "tables": dict(sorted(table_stats.items(), key=lambda item: int(item[0]))),
            "archive_inventory": archive_inventory,
            "parsed_table_count": len(table_counts),
            "unparsed_members": unparsed_members,
            "ignored_members": ignored_members,
        },
    )


def parse_huri(context: ParsingContext, output_root: Path) -> dict[str, Any]:
    cfg = dict(context.config["sources"]["huri"])
    pair_summary, pair_stats = _parse_pair_views(context, output_root, cfg)
    evidence_summary, participant_summary, feature_summary, mitab_stats = _parse_mitab(
        context, output_root, cfg
    )
    (
        supplement_summary,
        orf_summary,
        space_summary,
        contact_summary,
        fusion_summary,
        supplement_stats,
    ) = _parse_supplementary_tables(context, output_root, cfg)
    return {
        "source": "huri",
        "release": str(cfg["source_release"]),
        "parser_revision": "huri_v2_complete_workbook_coverage",
        "pair_assets": pair_stats,
        "mitab_assets": mitab_stats,
        "supplement": supplement_stats,
        "tables": {
            "source_pair_views": pair_summary,
            "evidence_records": evidence_summary,
            "participants": participant_summary,
            "participant_features": feature_summary,
            "supplementary_raw_tabular_records": supplement_summary,
            "huri_orf_mappings": orf_summary,
            "huri_space_membership": space_summary,
            "huri_structural_contact_annotations": contact_summary,
            "huri_fusion_interference": fusion_summary,
        },
    }
